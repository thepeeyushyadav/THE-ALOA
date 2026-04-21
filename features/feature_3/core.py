import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Global Variable
CURRENT_FILE_PATH = None
CURRENT_DF = None

def load_and_view_data(file_path):
    """Loads file and returns student list for display."""
    global CURRENT_FILE_PATH, CURRENT_DF
    
    if not os.path.exists(file_path):
        return None, "❌ Error: File not found."
    
    try:
        # Load Data
        df = pd.read_excel(file_path, dtype=str)
        df.columns = df.columns.str.strip() # Clean column names
        CURRENT_FILE_PATH = file_path
        CURRENT_DF = df
        
        # Validate Columns
        required = ['Enrollment NO.', 'Student Name']
        if not all(col in df.columns for col in required):
            return None, f"❌ Invalid Format: Missing required columns {required}."

        # Create Short List for Display
        student_list = []
        for index, row in df.iterrows():
            enroll = str(row['Enrollment NO.'])
            last_4 = enroll[-4:] # Extract last 4 digits
            name = row['Student Name']
            student_list.append(f"[{last_4}] {name}")
            
        return student_list, f"✅ Data Loaded: {len(student_list)} students identified."
        
    except Exception as e:
        return None, f"❌ System Error: {str(e)}"

def process_absentees(absent_inputs):
    """Processes user input to find absentees."""
    global CURRENT_DF
    if CURRENT_DF is None: return [], [], []

    absent_indices = []
    absent_names = []
    not_found = []

    tokens = absent_inputs.split()

    for token in tokens:
        token = token.strip()
        found = False
        
        for idx, row in CURRENT_DF.iterrows():
            enroll = str(row['Enrollment NO.'])
            name = str(row['Student Name'])
            
            if enroll.endswith(token) or token.lower() in name.lower():
                if idx not in absent_indices:
                    absent_indices.append(idx)
                    absent_names.append(f"{name} ({enroll[-4:]})")
                found = True
                break
        
        if not found:
            not_found.append(token)

    return absent_names, absent_indices, not_found

def save_final_attendance(absent_indices, attendance_date):
    """
    Saves attendance for the SPECIFIC DATE provided by user.
    """
    global CURRENT_FILE_PATH
    
    try:
        wb = load_workbook(CURRENT_FILE_PATH)
        ws = wb.active 
        
        # 1. Handle Date Header (Use the user-provided date)
        header_row = 1
        date_col_idx = None
        
        # Check if this specific date column already exists
        for cell in ws[header_row]:
            # Convert cell value to string to compare safely
            cell_val = str(cell.value).strip() if cell.value else ""
            if cell_val == attendance_date:
                date_col_idx = cell.column
                break
        
        # If not, create new column in 1st Row
        if date_col_idx is None:
            date_col_idx = ws.max_column + 1
            ws.cell(row=header_row, column=date_col_idx, value=attendance_date)

        # 2. Define Styles
        red_font = Font(color="FF0000", bold=True)
        green_font = Font(color="008000")
        center_align = Alignment(horizontal='center')

        # 3. Mark Attendance
        total_rows = len(CURRENT_DF)
        
        for i in range(total_rows):
            excel_row = i + 2
            cell = ws.cell(row=excel_row, column=date_col_idx)
            
            if i in absent_indices:
                cell.value = "ABSENT"
                cell.font = red_font
            else:
                cell.value = "P"
                cell.font = green_font
            
            cell.alignment = center_align

        wb.save(CURRENT_FILE_PATH)
        return f"✅ Success: Attendance for '{attendance_date}' saved successfully."
        
    except PermissionError:
        return "❌ Permission Error: The Excel file is currently open. Please close it and try again."
    except Exception as e:
        return f"❌ Write Error: {e}"